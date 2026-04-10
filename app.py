import streamlit as st
import pandas as pd
from rapidfuzz import process, fuzz
from io import BytesIO
import plotly.express as px
from openpyxl.styles import PatternFill

# PDF
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

st.set_page_config(page_title="Analizador PRO", page_icon="🚀", layout="wide")

st.title("🚀 Analizador de Coincidencias - SMART PRO")

st.write("Sube dos archivos de Excel y selecciona las hojas y columnas.")

archivo1 = st.file_uploader("📂 Archivo 1", type=["xlsx"])
archivo2 = st.file_uploader("📂 Archivo 2", type=["xlsx"])

if archivo1 and archivo2:

    excel1 = pd.ExcelFile(archivo1)
    excel2 = pd.ExcelFile(archivo2)

    hoja1 = st.selectbox("Hoja archivo 1", excel1.sheet_names)
    hoja2 = st.selectbox("Hoja archivo 2", excel2.sheet_names)

    if hoja1 and hoja2:

        df1 = pd.read_excel(excel1, sheet_name=hoja1)
        df2 = pd.read_excel(excel2, sheet_name=hoja2)

        col1 = st.selectbox("Columna archivo 1", df1.columns)
        col2 = st.selectbox("Columna archivo 2", df2.columns)

        if col1 and col2:

            umbral = st.slider("Umbral de similitud", 0, 100, 85)

            base1 = df1[col1].dropna().astype(str).str.lower().str.strip()
            base2 = df2[col2].dropna().astype(str).str.lower().str.strip()

            # 🔍 Duplicados
            duplicados_base1 = base1[base1.duplicated(keep=False)].unique()
            duplicados_base2 = base2[base2.duplicated(keep=False)].unique()

            # 🔄 Matching
            def emparejar_bases(base1, base2, threshold):
                emparejados = []
                base2_lista = base2.tolist()
                usados = set()

                for nombre1 in base1:
                    match = process.extractOne(
                        nombre1,
                        base2_lista,
                        scorer=fuzz.token_sort_ratio
                    )

                    if match and match[1] >= threshold and match[0] not in usados:
                        emparejados.append([nombre1, match[0], match[1], "Coincidencia"])
                        usados.add(match[0])
                    else:
                        emparejados.append([nombre1, None, 0, "Sin coincidencia"])

                for nombre2 in base2_lista:
                    if nombre2 not in usados:
                        emparejados.append([None, nombre2, 0, "Sin coincidencia"])

                return pd.DataFrame(
                    emparejados,
                    columns=[f"Base {col1}", f"Base {col2}", "Similitud (%)", "Estado"]
                )

            df_emparejados = emparejar_bases(base1, base2, umbral)

            # =========================
            # 📊 DASHBOARD
            # =========================
            st.write("## 📊 Dashboard")

            total = len(df_emparejados)
            coincidencias = df_emparejados[df_emparejados["Estado"] == "Coincidencia"].shape[0]
            no_coincidencias = total - coincidencias
            promedio_similitud = df_emparejados["Similitud (%)"].mean()

            colA, colB, colC, colD = st.columns(4)

            colA.metric("Total", total)
            colB.metric("Coincidencias", coincidencias)
            colC.metric("No coincidencias", no_coincidencias)
            colD.metric("Promedio", f"{promedio_similitud:.2f}%")

            # Gráficos
            fig1 = px.pie(df_emparejados, names="Estado", title="Distribución")
            st.plotly_chart(fig1, use_container_width=True)

            fig2 = px.histogram(df_emparejados, x="Similitud (%)", nbins=20, title="Similitud")
            st.plotly_chart(fig2, use_container_width=True)

            # Filtro
            filtro = st.selectbox("Filtrar", ["Todos", "Coincidencia", "Sin coincidencia"])

            if filtro != "Todos":
                df_filtrado = df_emparejados[df_emparejados["Estado"] == filtro]
            else:
                df_filtrado = df_emparejados

            st.dataframe(df_filtrado)

            # =========================
            # 🎨 EXCEL PRO
            # =========================
            def convertir_a_excel(df_main, dup1, dup2):
                output = BytesIO()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_main.to_excel(writer, sheet_name="Matching", index=False)
                    pd.DataFrame(dup1, columns=[col1]).to_excel(writer, sheet_name="Duplicados Base1", index=False)
                    pd.DataFrame(dup2, columns=[col2]).to_excel(writer, sheet_name="Duplicados Base2", index=False)

                    sheet = writer.sheets["Matching"]

                    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    col_estado = list(df_main.columns).index("Estado") + 1

                    for row in range(2, len(df_main) + 2):
                        estado = sheet.cell(row=row, column=col_estado).value
                        fill = verde if estado == "Coincidencia" else rojo

                        for col in range(1, len(df_main.columns) + 1):
                            sheet.cell(row=row, column=col).fill = fill

                return output.getvalue()

            excel_data = convertir_a_excel(df_emparejados, duplicados_base1, duplicados_base2)

            nombre_archivo = st.text_input("Nombre archivo Excel", "reporte-pro")

            st.download_button(
                "📥 Descargar Excel",
                excel_data,
                file_name=f"{nombre_archivo}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # =========================
            # 📄 PDF
            # =========================
            def generar_pdf(total, coincidencias, no_coincidencias, promedio):
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer)

                styles = getSampleStyleSheet()
                content = []

                content.append(Paragraph("Reporte de Coincidencias", styles["Title"]))
                content.append(Spacer(1, 12))

                content.append(Paragraph(f"Total: {total}", styles["Normal"]))
                content.append(Paragraph(f"Coincidencias: {coincidencias}", styles["Normal"]))
                content.append(Paragraph(f"No coincidencias: {no_coincidencias}", styles["Normal"]))
                content.append(Paragraph(f"Promedio similitud: {promedio:.2f}%", styles["Normal"]))

                doc.build(content)
                buffer.seek(0)
                return buffer

            pdf_file = generar_pdf(total, coincidencias, no_coincidencias, promedio_similitud)

            st.download_button(
                "📄 Descargar PDF",
                pdf_file,
                file_name="reporte_dashboard.pdf",
                mime="application/pdf"
            )
